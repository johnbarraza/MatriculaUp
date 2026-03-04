"""
Inspect SSU Excel and EFE PDF for data structure understanding.
Run: python scripts/inspect_ssu.py
"""
import pdfplumber
import openpyxl

EXCEL_PATH = r"pdfs/matricula/2026-1/EFEs/Sesiones SSU_2026-I(1).xlsx"
PDF_PATH = r"pdfs/matricula/2026-1/EFEs/Horarios-ofertados-en-matricula-2026-I_planes-antiguos.pdf"

# ── 1. INSPECT EXCEL ─────────────────────────────────────────────────────────
print("=" * 60)
print("EXCEL INSPECTION")
print("=" * 60)
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active
print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row}  Cols: {ws.max_column}")
print()

rows = list(ws.iter_rows(values_only=True))
print("First 10 rows (raw):")
for i, row in enumerate(rows[:10]):
    print(f"  Row {i+1}: {row}")

print()
print("Scanning for non-empty header rows and structure...")
# Find where the actual data starts
for i, row in enumerate(rows[:30]):
    if any(c is not None for c in row):
        print(f"  Row {i+1}: {row}")

print()
print("Sample data rows (first 20 with content):")
count = 0
for i, row in enumerate(rows):
    if i == 0:
        continue
    if any(c is not None for c in row):
        print(f"  Row {i+1}: {row}")
        count += 1
        if count >= 20:
            break

# ── 2. INSPECT PDF ────────────────────────────────────────────────────────────
print()
print("=" * 60)
print("PDF INSPECTION -- All pages/tables")
print("=" * 60)

with pdfplumber.open(PDF_PATH) as pdf:
    print(f"Total pages: {len(pdf.pages)}")
    for i, page in enumerate(pdf.pages):
        tables = page.find_tables()
        print(f"\n--- Page {i+1} ({len(tables)} tables) ---")
        for j, tbl in enumerate(tables):
            rows_raw = tbl.extract()
            print(f"  Table {j+1} ({len(rows_raw)} rows):")
            for rn, r in enumerate(rows_raw):
                print(f"    [{rn:02d}] {r}")
